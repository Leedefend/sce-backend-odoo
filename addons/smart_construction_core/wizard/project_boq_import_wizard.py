# -*- coding: utf-8 -*-
import base64
import csv
import io
import re

from odoo import fields, models
from odoo.exceptions import UserError
from odoo.tools import misc

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None


class ProjectBoqImportWizard(models.TransientModel):
    _name = "project.boq.import.wizard"
    _description = "工程量清单导入"

    BATCH_CREATE_SIZE = 500

    UOM_ALIAS_MAP = {
        "m2": ["㎡", "m²", "平米", "平方米", "平方"],
        "m3": ["立方", "立方米"],
        "item": ["项", "项（包干）", "项(包干)", "item"],
    }

    project_id = fields.Many2one(
        "project.project",
        string="项目",
        required=True,
    )
    section_type = fields.Selection(
        [
            ("building", "建筑"),
            ("installation", "安装/机电"),
            ("decoration", "装饰"),
            ("landscape", "景观"),
            ("other", "其他"),
        ],
        string="工程类别",
        help="若文件未识别到工程类别，使用此处的默认值。",
    )
    boq_category = fields.Selection(
        [
            ("boq", "分部分项清单"),
            ("unit_measure", "单价措施清单"),
            ("total_measure", "总价措施清单"),
            ("fee", "规费"),
            ("tax", "税金"),
        ],
        string="清单类别",
        default="boq",
        required=True,
        help="标识清单来源类别，避免分部分项与措施清单互相混淆。",
    )
    single_name = fields.Char("单项工程")
    unit_name = fields.Char("单位工程")
    source_type = fields.Selection(
        [
            ("tender", "招标清单"),
            ("contract", "合同清单"),
            ("settlement", "结算清单"),
        ],
        string="清单来源",
        default="contract",
    )
    version = fields.Char("版本号/批次", default="V1")
    clear_mode = fields.Selection(
        [
            ("append", "追加"),
            ("replace_project", "清空当前项目后导入"),
            ("replace_code", "按编码覆盖"),
        ],
        string="导入策略",
        default="append",
        required=True,
    )
    file = fields.Binary(string="导入文件", required=True)
    filename = fields.Char("文件名")
    log = fields.Text("导入日志", readonly=True)
    note = fields.Html(
        string="导入说明",
        readonly=True,
        default=lambda self: (
            "<ul>"
            "<li>同一编码在表中多次出现，将导入为多条清单行，并在工程结构中归入同一清单子目节点。</li>"
            "<li>若单位不存在，将自动规范化并创建新的计量单位。</li>"
            "<li>导入策略：追加 / 清空项目后导入 / 按编码覆盖。</li>"
            "</ul>"
        ),
    )

    def action_import(self):
        self.ensure_one()
        if not self.file:
            raise UserError("请先上传导入文件。")

        rows, created_uoms, skipped = self._parse_file()
        if not rows:
            raise UserError(
                "未找到可导入的清单数据：\n"
                "请确认文件中至少有一行同时包含名称列（清单名称/项目名称/汇总内容等）"
                "并且数量/单价/金额至少有一个为数字。"
            )

        Boq = self.env["project.boq.line"]
        created_count = 0
        updated_count = 0
        if self.clear_mode == "replace_project":
            Boq.search([("project_id", "=", self.project_id.id)]).unlink()
            created_count = self._batch_create(Boq, rows)
        elif self.clear_mode == "replace_code":
            for vals in rows:
                domain = [
                    ("project_id", "=", vals["project_id"]),
                    ("code", "=", vals.get("code")),
                    ("boq_category", "=", vals.get("boq_category", False)),
                    ("source_type", "=", vals.get("source_type")),
                    ("version", "=", vals.get("version")),
                ]
                existing = Boq.search(domain, limit=1)
                if existing:
                    existing.write(vals)
                    updated_count += 1
                else:
                    Boq.create(vals)
                    created_count += 1
        else:
            created_count = self._batch_create(Boq, rows)

        log_lines = []
        log_lines.append(f"成功导入 {created_count} 条，更新 {updated_count} 条。")
        if skipped:
            log_lines.append(f"跳过 {skipped} 行（空行/小计行/无数值行）。")
        if created_uoms:
            log_lines.append("自动创建计量单位：\n- " + "\n- ".join(sorted(created_uoms)))
            self.project_id.message_post(body=log_lines[-1])
        log_lines.append("如需刷新工程结构，请在项目中点击“生成工程结构”按钮。")
        self.log = "\n".join(log_lines)

        return {
            "type": "ir.actions.act_window",
            "res_model": "project.boq.line",
            "view_mode": "tree,form",
            "domain": [("project_id", "=", self.project_id.id)],
            "context": {"default_project_id": self.project_id.id},
            "target": "current",
        }

    # --- helpers ---

    def _parse_file(self):
        """Parse CSV/XLS/XLSX into vals list for project.boq.line."""
        data = base64.b64decode(self.file)
        filename = (self.filename or "").lower()

        # Excel：xlsx/xls
        if filename.endswith((".xlsx", ".xls")):
            return self._parse_excel(data, filename)

        # 默认按 UTF-8/GBK CSV 解析
        content = self._read_as_csv(data)
        return self._parse_csv_content(content)


    def _parse_csv_content(self, content):
        reader = csv.reader(io.StringIO(content))
        rows_data = list(reader)
        if not rows_data:
            raise UserError("导入文件没有数据，请检查。")
        # 头部探测：使用第一行作为表头
        headers = [str(h or "").strip() for h in rows_data[0]]
        data_rows = rows_data[1:]
        col_map = self._prepare_col_map(headers)
        rows, created_uoms, skipped = self._build_rows_from_iter(
            data_rows, col_map, strict_numeric=True,
            default_single=self.single_name, default_unit=self.unit_name,
            boq_category=self.boq_category,
        )
        if not rows:
            rows, created_uoms, skipped = self._build_rows_from_iter(
                data_rows, col_map, strict_numeric=False,
                default_single=self.single_name, default_unit=self.unit_name,
                boq_category=self.boq_category,
            )
        return rows, created_uoms, skipped

    def _parse_excel(self, data, filename):
        col_map_cfg = self._col_map_cfg()
        rows_all = []
        created_uoms_all = set()
        skipped_all = 0

        if filename.endswith(".xlsx"):
            if not openpyxl:
                raise UserError("服务器缺少 openpyxl，无法解析 XLSX，请安装依赖或改用 CSV。")
            book = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            for idx, sheet in enumerate(book.worksheets, start=1):
                if not self._is_supported_sheet(sheet.title or ""):
                    continue
                header_info = self._extract_excel_header(sheet)
                if not header_info:
                    continue
                headers, data_start_row, header_row_idx = header_info
                parsed_single, parsed_unit, parsed_major = self._parse_engineering_header_excel(
                    sheet, limit=max(5, header_row_idx)
                )
                default_single = self.single_name or parsed_single
                default_unit = self.unit_name or parsed_unit
                section_type = self.section_type or self._map_major_to_section_type(parsed_major) or self._guess_section_type(sheet.title or "")
                category = self._detect_boq_category(sheet.title or "") or self.boq_category
                col_map = self._prepare_col_map(headers, col_map_cfg)
                data_rows = [list(row) for row in sheet.iter_rows(min_row=data_start_row, values_only=True)]
                rows, created_uoms, skipped = self._build_rows_from_iter(
                    data_rows, col_map, section_type, strict_numeric=True,
                    default_single=default_single, default_unit=default_unit,
                    major_name=parsed_major, sheet_index=idx, sheet_name=sheet.title or "",
                    boq_category=category,
                )
                if not rows:
                    rows, created_uoms, skipped = self._build_rows_from_iter(
                        data_rows, col_map, section_type, strict_numeric=False,
                        default_single=default_single, default_unit=default_unit,
                        major_name=parsed_major, sheet_index=idx, sheet_name=sheet.title or "",
                        boq_category=category,
                    )
                rows_all.extend(rows)
                created_uoms_all.update(created_uoms)
                skipped_all += skipped
            return rows_all, created_uoms_all, skipped_all

        if filename.endswith(".xls"):
            if not xlrd:
                raise UserError("服务器缺少 xlrd，无法解析 XLS，请安装依赖或改用 CSV。")
            book = xlrd.open_workbook(file_contents=data)
            for idx, sheet in enumerate(book.sheets(), start=1):
                if sheet.nrows < 1:
                    continue
                if not self._is_supported_sheet(sheet.name or ""):
                    continue
                headers, data_start_row, header_row_idx = self._extract_xls_header(sheet)
                if not headers:
                    continue
                parsed_single, parsed_unit, parsed_major = self._parse_engineering_header_xls(
                    sheet, limit=max(5, header_row_idx)
                )
                default_single = self.single_name or parsed_single
                default_unit = self.unit_name or parsed_unit
                section_type = self.section_type or self._map_major_to_section_type(parsed_major) or self._guess_section_type(sheet.name or "")
                category = self._detect_boq_category(sheet.name or "") or self.boq_category
                col_map = self._prepare_col_map(headers, col_map_cfg)
                data_rows = [
                    [sheet.cell_value(r, c) for c in range(sheet.ncols)]
                    for r in range(data_start_row, sheet.nrows)
                ]
                rows, created_uoms, skipped = self._build_rows_from_iter(
                    data_rows, col_map, section_type, strict_numeric=True,
                    default_single=default_single, default_unit=default_unit,
                    major_name=parsed_major, sheet_index=idx, sheet_name=sheet.name or "",
                    boq_category=category,
                )
                if not rows:
                    rows, created_uoms, skipped = self._build_rows_from_iter(
                        data_rows, col_map, section_type, strict_numeric=False,
                        default_single=default_single, default_unit=default_unit,
                        major_name=parsed_major, sheet_index=idx, sheet_name=sheet.name or "",
                        boq_category=category,
                    )
                rows_all.extend(rows)
                created_uoms_all.update(created_uoms)
                skipped_all += skipped
            return rows_all, created_uoms_all, skipped_all

        # 回退：尝试按 UTF-8/GBK 解码 CSV
        return self._parse_csv_bytes(data), set(), 0

    def _col_map_cfg(self):
        return {
            "code": ["清单编码", "编码", "code"],
            "name": ["清单名称", "名称", "name", "项目名称", "汇总内容"],
            "spec": ["规格", "规格型号", "项目特征", "项目特征描述"],
            "uom_id": ["单位", "uom"],
            "quantity": ["工程量", "数量", "qty"],
            "price": ["单价", "price"],
            # “金额（元）”多见于总标题，不直接当金额列匹配
            "amount": ["合价", "合计", "amount", "金额", "金额元"],
            "cost_item_id": ["成本项", "成本科目"],
            "remark": ["备注", "说明"],
        }

    def _prepare_col_map(self, headers, col_map_cfg=None):
        col_map_cfg = col_map_cfg or self._col_map_cfg()
        col_map = {}
        for idx, title in enumerate(headers):
            title_norm = self._normalize_header(title)
            for field, aliases in col_map_cfg.items():
                matched = False
                for alias in aliases:
                    alias_norm = self._normalize_header(alias)
                    if title_norm == alias_norm or title_norm.endswith(alias_norm) or alias_norm in title_norm:
                        matched = True
                        break
                if matched and field not in col_map:
                    col_map[field] = idx
                    break
        if "name" not in col_map:
            # 兜底：首列作为名称
            if headers:
                col_map["name"] = 0
            else:
                raise UserError("模板中至少需要包含 “清单名称” 列。")
        # 若识别到工程量列，按相对位置推断单价/合价（常见 F.1.1 结构：工程量右一列=单价，右二列=合价）
        qty_idx = col_map.get("quantity")
        if qty_idx is not None:
            if "price" not in col_map and qty_idx + 1 < len(headers):
                col_map["price"] = qty_idx + 1
            if "amount" not in col_map and qty_idx + 2 < len(headers):
                col_map["amount"] = qty_idx + 2
        return col_map

    def _find_header_in_sheet(self, row_iter):
        """Deprecated: use _extract_excel_header/_extract_xls_header."""
        return [], 0

    # ----------------- Excel helpers -----------------
    def _parse_engineering_header_excel(self, sheet, limit=5):
        """
        解析表头中的“工程名称：项目\\单位【专业】”
        返回 (single_name, unit_name, major_name)
        """
        for row in sheet.iter_rows(min_row=1, max_row=limit, values_only=True):
            for val in row:
                if not val:
                    continue
                text = str(val)
                if "工程名称" not in text:
                    continue
                text = text.split("工程名称", 1)[-1]
                text = text.lstrip("：:").strip()
                parts = text.split("\\")
                single = parts[0].strip() if parts else ""
                tail = parts[1].strip() if len(parts) >= 2 else ""
                unit = tail
                major = ""
                if "【" in tail and "】" in tail:
                    before = tail.split("【", 1)[0]
                    inner = tail.split("【", 1)[1].split("】", 1)[0]
                    unit = before.strip()
                    major = inner.strip()
                return single, unit, major
        return "", "", ""

    def _parse_engineering_header_xls(self, sheet, limit=5):
        for r in range(min(limit, sheet.nrows)):
            for c in range(sheet.ncols):
                val = sheet.cell_value(r, c)
                if not val:
                    continue
                text = str(val)
                if "工程名称" not in text:
                    continue
                text = text.split("工程名称", 1)[-1]
                text = text.lstrip("：:").strip()
                parts = text.split("\\")
                single = parts[0].strip() if parts else ""
                tail = parts[1].strip() if len(parts) >= 2 else ""
                unit = tail
                major = ""
                if "【" in tail and "】" in tail:
                    before = tail.split("【", 1)[0]
                    inner = tail.split("【", 1)[1].split("】", 1)[0]
                    unit = before.strip()
                    major = inner.strip()
                return single, unit, major
        return "", "", ""

    def _extract_excel_header(self, sheet, header_rows=3, scan_rows=8):
        """处理多行表头+合并单元格，返回(扁平列名列表, 数据起始行号)"""
        merge_map = {}
        try:
            merge_ranges = sheet.merged_cells
        except Exception:
            merge_ranges = None
        if merge_ranges:
            ranges = getattr(merge_ranges, "ranges", merge_ranges)
            try:
                for m in ranges:
                    min_row, min_col, max_row, max_col = m.min_row, m.min_col, m.max_row, m.max_col
                    for r in range(min_row, max_row + 1):
                        for c in range(min_col, max_col + 1):
                            merge_map[(r, c)] = (min_row, min_col)
            except Exception:
                merge_map = {}

        def cell_val(r, c):
            key = merge_map.get((r, c))
            if key:
                r, c = key
            return sheet.cell(row=r, column=c).value

        max_col = sheet.max_column or 0
        header_row_idx = 0
        best_hits = 0
        keywords = ["编码", "项目编码", "清单编码", "特征", "工程量", "综合单价", "合价", "计量单位"]
        for idx in range(1, min(scan_rows, sheet.max_row or 0) + 1):
            row_vals = [str(cell_val(idx, c) or "").strip() for c in range(1, max_col + 1)]
            hits = sum(1 for v in row_vals if any(k in v for k in keywords))
            if hits > best_hits:
                best_hits = hits
                header_row_idx = idx
        if not header_row_idx:
            return None

        header_rows_vals = []
        for r in range(header_row_idx, min(header_row_idx + header_rows, (sheet.max_row or 0) + 1)):
            row_vals = []
            for c in range(1, max_col + 1):
                row_vals.append(str(cell_val(r, c) or "").strip())
            header_rows_vals.append(row_vals)

        # 纵向拼接列名
        flat_headers = []
        for c in range(max_col):
            parts = []
            for r in range(len(header_rows_vals)):
                v = header_rows_vals[r][c]
                if v:
                    parts.append(v)
            flat_headers.append(" - ".join(parts) if parts else "")

        data_start = header_row_idx + header_rows
        return flat_headers, data_start, header_row_idx

    def _extract_xls_header(self, sheet, header_rows=3, scan_rows=8):
        max_col = sheet.ncols
        header_row_idx = 0
        best_hits = 0
        keywords = ["编码", "项目编码", "清单编码", "特征", "工程量", "综合单价", "合价", "计量单位"]
        for idx in range(min(scan_rows, sheet.nrows)):
            row_vals = [str(sheet.cell_value(idx, c) or "").strip() for c in range(max_col)]
            hits = sum(1 for v in row_vals if any(k in v for k in keywords))
            if hits > best_hits:
                best_hits = hits
                header_row_idx = idx
        if max_col == 0:
            return None, 0
        header_rows_vals = []
        for r in range(header_row_idx, min(header_row_idx + header_rows, sheet.nrows)):
            row_vals = [str(sheet.cell_value(r, c) or "").strip() for c in range(max_col)]
            header_rows_vals.append(row_vals)
        flat_headers = []
        for c in range(max_col):
            parts = [row_vals[c] for row_vals in header_rows_vals if row_vals[c]]
            flat_headers.append(" - ".join(parts) if parts else "")
        data_start = header_row_idx + header_rows
        return flat_headers, data_start, header_row_idx

    def _build_rows_from_iter(
        self, row_iter, col_map, section_type=None, strict_numeric=True,
        default_single=None, default_unit=None,
        major_name=None, sheet_index=None, sheet_name=None,
        boq_category=None,
    ):
        Uom = self.env["uom.uom"]
        Dict, dict_domain_key = self._get_dictionary_model()

        rows = []
        uom_cache = {}
        cost_item_cache = {}
        created_uoms = set()
        skipped_rows = 0
        current_division = None

        def _default_uom_category():
            """选用通用“单位”类别，若缺失则取任一类别兜底。"""
            category = self.env.ref("uom.product_uom_categ_unit", raise_if_not_found=False)
            if not category:
                category = self.env["uom.category"].search([], limit=1)
            return category

        for row in row_iter:
            # 小工具：按字段名取这一行对应列的值
            def get(field):
                idx = col_map.get(field)
                if idx is None or idx >= len(row):
                    return ""
                return row[idx] if not isinstance(row, dict) else row.get(idx)

            name = str(get("name") or "").strip()
            code = str(get("code") or "").strip()
            if not (name or code):
                skipped_rows += 1
                continue

            # 分部标题行：项目编码空、名称有值（且非数字） -> 记录当前分部，跳过本行
            if not code and name and not self._is_number(name):
                # 小计/合计类行作为分部标题会污染分部名称，直接跳过
                lower_name = name.lower()
                if any(key in lower_name for key in ["合计", "小计", "本页", "本表"]):
                    skipped_rows += 1
                    continue
                current_division = name
                continue

            vals = {
                "project_id": self.project_id.id,
                "name": name,
                "section_type": self.section_type or section_type or False,
                "single_name": default_single or self.single_name or False,
                "unit_name": default_unit or self.unit_name or False,
                "major_name": major_name or False,
                "sheet_index": sheet_index,
                "sheet_name": sheet_name,
                "source_type": self.source_type,
                "version": self.version,
                "boq_category": boq_category or self.boq_category or "boq",
            }
            if section_type and not vals["section_type"]:
                vals["section_type"] = section_type

            spec = str(get("spec") or "").strip()
            remark = str(get("remark") or "").strip()
            qty = get("quantity")
            price = get("price")
            amount_val = get("amount")

            if code:
                vals["code"] = code
            if spec:
                vals["spec"] = spec
            # 记录当前分部名称，便于后续 WBS 直接使用（无需再从 remark 里解析）
            if current_division:
                vals["division_name"] = current_division
            if remark or current_division:
                prefix = f"[分部]{current_division}" if current_division else ""
                vals["remark"] = f"{prefix} {remark}".strip() if (prefix or remark) else False

            vals["quantity"] = self._to_float(qty)
            vals["price"] = self._to_float(price)
            vals["amount"] = self._to_float(amount_val)

            # 若数量/单价/合价均为0，则视为标题/小计行跳过
            if strict_numeric:
                if not any(
                    [
                        self._is_number(qty),
                        self._is_number(price),
                        self._is_number(amount_val),
                    ]
                ):
                    skipped_rows += 1
                    continue

            # 常见小计/合计行过滤
            lower_name = (name or "").lower()
            if any(key in lower_name for key in ["合计", "小计", "本页", "本表"]):
                skipped_rows += 1
                continue

            # ===== 计量单位处理 =====
            uom = False
            uom_name = str(get("uom_id") or "").strip()
            if uom_name:
                norm_name = self._normalize_uom_name(uom_name)
                canonical = self._canonical_uom(norm_name)
                search_key = canonical or norm_name or uom_name

                uom = uom_cache.get(search_key)
                create_name = None

                if uom is None:
                    # 先按规范名找
                    uom = Uom.search([("name", "=", search_key)], limit=1)
                    # 再按原始名兜底
                    if not uom and uom_name != search_key:
                        uom = Uom.search([("name", "=", uom_name)], limit=1)

                    if not uom:
                        category = _default_uom_category()
                        if not category:
                            raise UserError(
                                "未找到计量单位类别，无法自动创建单位，请先在系统中创建一个计量单位类别。"
                            )
                        create_name = search_key
                        ref_uom = Uom.search(
                            [
                                ("category_id", "=", category.id),
                                ("uom_type", "=", "reference"),
                            ],
                            limit=1,
                        )
                        uom_vals = {
                            "name": create_name,
                            "category_id": category.id,
                            "factor": 1.0,
                            "factor_inv": 1.0,
                            "rounding": 0.0001,
                            "active": True,
                        }
                        # 如果类别已有参照单位，则新建等效单位用 smaller 并保持 factor=1
                        if ref_uom:
                            uom_vals["uom_type"] = "smaller"
                            uom_vals["factor"] = 1.0
                            uom_vals["factor_inv"] = 1.0
                        else:
                            uom_vals["uom_type"] = "reference"
                        uom = Uom.create(uom_vals)
                        if create_name:
                            created_uoms.add(create_name)

                if uom:
                    uom_cache[search_key] = uom

            # 若仍未找到单位，使用“单位”类别的参照单位兜底，避免必填校验失败
            if not uom:
                category = _default_uom_category()
                if category:
                    uom = Uom.search(
                        [
                            ("category_id", "=", category.id),
                            ("uom_type", "=", "reference"),
                        ],
                        limit=1,
                    )
                    if not uom:
                        # 创建一个通用参照单位“项”作为兜底
                        uom = Uom.create(
                            {
                                "name": "项",
                                "category_id": category.id,
                                "uom_type": "reference",
                                "factor": 1.0,
                                "factor_inv": 1.0,
                                "rounding": 0.0001,
                                "active": True,
                            }
                        )
                        created_uoms.add("项")

            vals["uom_id"] = uom.id if uom else False

            # ===== 成本项字典匹配 =====
            cost_item_name = str(get("cost_item_id") or "").strip()
            if cost_item_name and Dict:
                cost_item = cost_item_cache.get(cost_item_name)
                if cost_item is None:
                    if isinstance(dict_domain_key, (list, tuple)):
                        domain = list(dict_domain_key)
                    else:
                        domain = [(dict_domain_key, "=", "cost_item")]
                    domain.append(("name", "=", cost_item_name))
                    cost_item = Dict.search(domain, limit=1)
                    cost_item_cache[cost_item_name] = cost_item
                vals["cost_item_id"] = cost_item.id or False

            rows.append(vals)

        return rows, created_uoms, skipped_rows

    def _read_as_csv(self, data_bytes):
        """Return CSV string from raw bytes."""
        return self._parse_csv_bytes(data_bytes)

    def _parse_csv_bytes(self, data_bytes):
        """Try utf-8, then gbk."""
        for encoding in ("utf-8", "gbk"):
            try:
                return data_bytes.decode(encoding)
            except Exception:
                continue
        raise UserError("无法解码导入文件，请确认使用 UTF-8 或 GBK 编码。")

    @staticmethod
    def _guess_section_type(sheet_title):
        title = (sheet_title or "").lower()
        mapping = {
            "build": "building",
            "建筑": "building",
            "机电": "installation",
            "安装": "installation",
            "elect": "installation",
            "装饰": "decoration",
            "decoration": "decoration",
            "景观": "landscape",
            "landscape": "landscape",
        }
        for key, val in mapping.items():
            if key in title:
                return val
        return False

    @staticmethod
    def _is_supported_sheet(title):
        """仅处理常见清单表，跳过封面/汇总等无效 sheet，匹配更宽。"""
        text = (title or "").replace(" ", "")
        keywords = [
            # 分部分项
            "分部分项工程清单", "分部分项工程量清单",
            # 单价措施
            "单价措施项目清单", "单价措施", "措施项目清单",
            # 总价措施
            "总价措施项目清单", "总价措施",
        ]
        return any(key in text for key in keywords)

    @staticmethod
    def _map_major_to_section_type(major_name):
        """根据专业名称映射工程类别（section_type）"""
        text = (major_name or "").lower()
        mapping = {
            "装饰": "decoration",
            "装修": "decoration",
            "建筑": "building",
            "土建": "building",
            "给排水": "installation",
            "暖通": "installation",
            "电气": "installation",
            "强电": "installation",
            "弱电": "installation",
            "机电": "installation",
            "消防": "installation",
            "安装": "installation",
            "景观": "landscape",
        }
        for key, val in mapping.items():
            if key in text:
                return val
        return False

    @staticmethod
    def _detect_boq_category(sheet_title):
        """根据 sheet 名推断清单类别：分部分项/单价措施/总价措施。"""
        title = (sheet_title or "").lower()
        if any(k in title for k in ["单价措施", "措施清单", "措施项目"]):
            return "unit_measure"
        if any(k in title for k in ["总价措施"]):
            return "total_measure"
        return "boq"

    @staticmethod
    def _normalize_header(title):
        text = str(title or "").strip()
        text = re.sub(r"\s+", "", text)
        return text.lower()

    @staticmethod


    def _is_number(val):
        try:
            if isinstance(val, str):
                cleaned = ProjectBoqImportWizard._clean_number_str(val)
                if cleaned in ("", "-", "--"):
                    return False
                float(cleaned)
            else:
                float(val)
            return True
        except Exception:
            return False

    @staticmethod
    def _to_float(val):
        try:
            if isinstance(val, str):
                cleaned = ProjectBoqImportWizard._clean_number_str(val)
                return float(cleaned or 0.0)
            return float(val or 0.0)
        except Exception:
            return 0.0

    def _get_dictionary_model(self):
        """返回可用的字典模型及类型字段键。"""
        dict_model = "project.dictionary" if "project.dictionary" in self.env.registry else "sc.dictionary"
        if dict_model not in self.env.registry:
            return None, None
        Dict = self.env[dict_model]
        fields_map = Dict._fields
        if "type" in fields_map:
            return Dict, "type"
        if "type_id" in fields_map:
            return Dict, "type_id.code"
        return Dict, "type"

    @staticmethod
    def _clean_number_str(text):
        """Remove common thousand separators and spaces."""
        cleaned = str(text or "")
        cleaned = cleaned.replace(",", "").replace("，", "").replace(" ", "").strip()
        return cleaned

    # --- UoM helpers ---
    def _normalize_uom_name(self, name):
        """基本规范化：去空格、全角转半角、小写。"""
        text = misc.ustr(name or "").strip()
        text = re.sub(r"\s+", "", text)
        res = []
        for ch in text:
            code = ord(ch)
            if 0xFF01 <= code <= 0xFF5E:
                code -= 0xfee0
                ch = chr(code)
            res.append(ch)
        return "".join(res).lower()

    def _canonical_uom(self, norm_name):
        """根据别名映射返回规范名，否则返回原名。"""
        for main, aliases in self.UOM_ALIAS_MAP.items():
            if norm_name == main:
                return main
            for alias in aliases:
                if norm_name == self._normalize_uom_name(alias):
                    return main
        return norm_name

    def _batch_create(self, model, vals_list):
        """批量创建，避免一次性巨大列表占用内存/锁时间过长。"""
        if not vals_list:
            return 0
        size = self.BATCH_CREATE_SIZE or 500
        created = 0
        for start in range(0, len(vals_list), size):
            chunk = vals_list[start : start + size]
            model.create(chunk)
            created += len(chunk)
        return created
